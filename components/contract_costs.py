import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots

def render_contract_costs():
    """Render the contract costs management interface"""
    
    st.markdown("## 💰 Contract Costs & P&L Management")
    st.markdown("Manage ODC, indirect costs, and track contract profitability")
    
    # Contract Costs Tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "📊 Cost Overview", 
        "💼 ODC Management", 
        "🏢 Indirect Costs", 
        "📈 P&L Analysis", 
        "📥 Import Data"
    ])
    
    with tab1:
        render_cost_overview()
    
    with tab2:
        render_odc_management()
    
    with tab3:
        render_indirect_costs()
    
    with tab4:
        render_pl_analysis()
    
    with tab5:
        render_cost_import()

def render_cost_overview():
    """Render cost overview dashboard"""
    
    st.markdown("### 📊 Contract Cost Overview")
    
    # Get cost data
    cost_data = st.session_state.data_manager.get_cost_summary()
    
    # Key Metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_labor = cost_data.get('total_labor', 0)
        st.metric(
            "👥 Total Labor Costs",
            f"${total_labor:,.0f}",
            delta=f"${cost_data.get('labor_variance', 0):,.0f}"
        )
    
    with col2:
        total_odc = cost_data.get('total_odc', 0)
        st.metric(
            "💼 Total ODC",
            f"${total_odc:,.0f}",
            delta=f"${cost_data.get('odc_variance', 0):,.0f}"
        )
    
    with col3:
        total_indirect = cost_data.get('total_indirect', 0)
        st.metric(
            "🏢 Total Indirect",
            f"${total_indirect:,.0f}",
            delta=f"${cost_data.get('indirect_variance', 0):,.0f}"
        )
    
    with col4:
        contract_value = cost_data.get('contract_value', 0)
        total_costs = total_labor + total_odc + total_indirect
        profit_loss = contract_value - total_costs
        
        st.metric(
            "📈 Profit/Loss",
            f"${profit_loss:,.0f}",
            delta=f"{(profit_loss/contract_value*100 if contract_value > 0 else 0):.1f}%"
        )
    
    # Cost Breakdown Charts
    st.markdown("---")
    col1, col2 = st.columns(2)
    
    with col1:
        # Cost breakdown pie chart
        if total_costs > 0:
            cost_categories = ['Labor', 'ODC', 'Indirect']
            cost_values = [total_labor, total_odc, total_indirect]
            
            fig_pie = px.pie(
                values=cost_values,
                names=cost_categories,
                title="Cost Breakdown by Category",
                color_discrete_sequence=['#2E86AB', '#A23B72', '#F18F01']
            )
            st.plotly_chart(fig_pie, use_container_width=True)
    
    with col2:
        # Monthly cost trend
        monthly_data = st.session_state.data_manager.get_monthly_costs()
        if not monthly_data.empty:
            fig_trend = px.line(
                monthly_data,
                x='month',
                y=['labor', 'odc', 'indirect'],
                title="Monthly Cost Trends",
                color_discrete_sequence=['#2E86AB', '#A23B72', '#F18F01']
            )
            fig_trend.update_layout(
                xaxis_title="Month",
                yaxis_title="Cost ($)",
                legend_title="Cost Type"
            )
            st.plotly_chart(fig_trend, use_container_width=True)
    
    # Contract Performance Summary
    st.markdown("---")
    st.markdown("### 📋 Contract Performance Summary")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        budget_utilization = (total_costs / contract_value * 100) if contract_value > 0 else 0
        st.metric("Budget Utilization", f"{budget_utilization:.1f}%")
        
        # Progress bar
        progress_color = "green" if budget_utilization <= 90 else "orange" if budget_utilization <= 100 else "red"
        st.progress(min(budget_utilization/100, 1.0))
    
    with col2:
        margin_percent = (profit_loss / contract_value * 100) if contract_value > 0 else 0
        st.metric("Profit Margin", f"{margin_percent:.1f}%")
    
    with col3:
        burn_rate = st.session_state.data_manager.get_monthly_burn_rate()
        st.metric("Monthly Burn Rate", f"${burn_rate:,.0f}")

def render_odc_management():
    """Render ODC (Other Direct Costs) management"""
    
    st.markdown("### 💼 Other Direct Costs (ODC) Management")
    
    # Add new ODC item form
    with st.expander("➕ Add New ODC Item", expanded=False):
        with st.form("add_odc_form"):
            col1, col2 = st.columns(2)
            
            with col1:
                odc_category = st.selectbox(
                    "ODC Category",
                    ["Travel", "Equipment", "Software", "Training", "Subcontractor", "Materials", "Other"]
                )
                description = st.text_input("Description", placeholder="Brief description of the cost")
                vendor = st.text_input("Vendor/Supplier", placeholder="Vendor or supplier name")
            
            with col2:
                amount = st.number_input("Amount ($)", min_value=0.0, step=100.0)
                date = st.date_input("Date", value=datetime.now())
                status = st.selectbox("Status", ["Planned", "Committed", "Invoiced", "Paid"])
            
            notes = st.text_area("Notes", placeholder="Additional notes or details")
            
            if st.form_submit_button("➕ Add ODC Item", type="primary"):
                if description and amount > 0:
                    odc_data = {
                        'category': odc_category,
                        'description': description,
                        'vendor': vendor,
                        'amount': amount,
                        'date': date,
                        'status': status,
                        'notes': notes
                    }
                    
                    st.session_state.data_manager.add_odc_item(odc_data)
                    st.success(f"✅ ODC item '{description}' added successfully!")
                    st.rerun()
                else:
                    st.error("❌ Please fill in description and amount")
    
    # Current ODC items
    odc_df = st.session_state.data_manager.get_odc_df()
    
    if odc_df.empty:
        st.info("💼 No ODC items found. Add some using the form above.")
    else:
        st.markdown(f"### 📋 Current ODC Items ({len(odc_df)} items)")
        
        # Filters
        col1, col2, col3 = st.columns(3)
        
        with col1:
            categories = ['All'] + list(odc_df['category'].unique())
            selected_category = st.selectbox("Filter by Category", categories)
        
        with col2:
            statuses = ['All'] + list(odc_df['status'].unique())
            selected_status = st.selectbox("Filter by Status", statuses)
        
        with col3:
            search_term = st.text_input("🔍 Search ODC", placeholder="Search description or vendor...")
        
        # Apply filters
        filtered_df = odc_df.copy()
        
        if selected_category != 'All':
            filtered_df = filtered_df[filtered_df['category'] == selected_category]
        
        if selected_status != 'All':
            filtered_df = filtered_df[filtered_df['status'] == selected_status]
        
        if search_term:
            mask = (
                filtered_df['description'].str.contains(search_term, case=False, na=False) |
                filtered_df['vendor'].str.contains(search_term, case=False, na=False)
            )
            filtered_df = filtered_df[mask]
        
        # Display ODC items
        if not filtered_df.empty:
            st.dataframe(
                filtered_df[['date', 'category', 'description', 'vendor', 'amount', 'status']],
                use_container_width=True
            )
            
            # ODC Summary
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_odc = filtered_df['amount'].sum()
                st.metric("Total ODC", f"${total_odc:,.0f}")
            
            with col2:
                planned = filtered_df[filtered_df['status'] == 'Planned']['amount'].sum()
                st.metric("Planned", f"${planned:,.0f}")
            
            with col3:
                committed = filtered_df[filtered_df['status'] == 'Committed']['amount'].sum()
                st.metric("Committed", f"${committed:,.0f}")
            
            with col4:
                paid = filtered_df[filtered_df['status'] == 'Paid']['amount'].sum()
                st.metric("Paid", f"${paid:,.0f}")
        else:
            st.info("No ODC items match the current filters.")

def render_indirect_costs():
    """Render indirect costs management"""
    
    st.markdown("### 🏢 Indirect Costs Management")
    st.markdown("Manage Fringe, Overhead, and G&A costs by period")
    
    # Indirect cost configuration
    with st.expander("⚙️ Configure Indirect Cost Rates", expanded=False):
        st.markdown("Set your standard indirect cost rates")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            fringe_rate = st.number_input("Fringe Rate (%)", min_value=0.0, max_value=100.0, value=25.0, step=0.1)
        
        with col2:
            overhead_rate = st.number_input("Overhead Rate (%)", min_value=0.0, max_value=200.0, value=45.0, step=0.1)
        
        with col3:
            ga_rate = st.number_input("G&A Rate (%)", min_value=0.0, max_value=100.0, value=15.0, step=0.1)
        
        if st.button("💾 Save Rates"):
            rates = {
                'fringe_rate': fringe_rate,
                'overhead_rate': overhead_rate,
                'ga_rate': ga_rate
            }
            st.session_state.data_manager.update_indirect_rates(rates)
            st.success("✅ Indirect cost rates updated!")
    
    # Monthly indirect costs input
    st.markdown("### 📅 Monthly Indirect Costs")
    
    # Get current indirect costs data
    indirect_df = st.session_state.data_manager.get_indirect_costs_df()
    
    # Period selector
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Year", [2024, 2025, 2026])
    with col2:
        month = st.selectbox("Month", list(range(1, 13)), format_func=lambda x: datetime(2024, x, 1).strftime('%B'))
    
    period_key = f"{year}-{month:02d}"
    
    # Get existing data for this period
    existing_data = st.session_state.data_manager.get_indirect_costs_for_period(period_key)
    
    # Input form for this period
    with st.form(f"indirect_costs_{period_key}"):
        st.markdown(f"**Indirect Costs for {datetime(year, month, 1).strftime('%B %Y')}**")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            fringe = st.number_input(
                "Fringe ($)",
                min_value=0.0,
                value=float(existing_data.get('fringe', 0)),
                step=100.0
            )
        
        with col2:
            overhead = st.number_input(
                "Overhead ($)",
                min_value=0.0,
                value=float(existing_data.get('overhead', 0)),
                step=100.0
            )
        
        with col3:
            ga = st.number_input(
                "G&A ($)",
                min_value=0.0,
                value=float(existing_data.get('ga', 0)),
                step=100.0
            )
        
        if st.form_submit_button("💾 Save Period Costs", type="primary"):
            period_data = {
                'period': period_key,
                'fringe': fringe,
                'overhead': overhead,
                'ga': ga,
                'total': fringe + overhead + ga
            }
            
            st.session_state.data_manager.update_indirect_costs_period(period_data)
            st.success(f"✅ Indirect costs for {datetime(year, month, 1).strftime('%B %Y')} updated!")
            st.rerun()
    
    # Display indirect costs summary
    if not indirect_df.empty:
        st.markdown("### 📊 Indirect Costs Summary")
        
        # Summary metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_fringe = indirect_df['fringe'].sum()
            st.metric("Total Fringe", f"${total_fringe:,.0f}")
        
        with col2:
            total_overhead = indirect_df['overhead'].sum()
            st.metric("Total Overhead", f"${total_overhead:,.0f}")
        
        with col3:
            total_ga = indirect_df['ga'].sum()
            st.metric("Total G&A", f"${total_ga:,.0f}")
        
        with col4:
            total_indirect = indirect_df['total'].sum()
            st.metric("Total Indirect", f"${total_indirect:,.0f}")
        
        # Indirect costs trend chart
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=indirect_df['period'],
            y=indirect_df['fringe'],
            mode='lines+markers',
            name='Fringe',
            line=dict(color='#2E86AB')
        ))
        
        fig.add_trace(go.Scatter(
            x=indirect_df['period'],
            y=indirect_df['overhead'],
            mode='lines+markers',
            name='Overhead',
            line=dict(color='#A23B72')
        ))
        
        fig.add_trace(go.Scatter(
            x=indirect_df['period'],
            y=indirect_df['ga'],
            mode='lines+markers',
            name='G&A',
            line=dict(color='#F18F01')
        ))
        
        fig.update_layout(
            title="Indirect Costs by Period",
            xaxis_title="Period",
            yaxis_title="Cost ($)",
            hovermode='x unified'
        )
        
        st.plotly_chart(fig, use_container_width=True)

def render_pl_analysis():
    """Render profit and loss analysis"""
    
    st.markdown("### 📈 Profit & Loss Analysis")
    
    # Contract settings
    with st.expander("⚙️ Contract Settings", expanded=False):
        col1, col2 = st.columns(2)
        
        with col1:
            contract_value = st.number_input(
                "Contract Value ($)",
                min_value=0.0,
                value=float(st.session_state.data_manager.get_contract_value()),
                step=10000.0
            )
        
        with col2:
            contract_start = st.date_input(
                "Contract Start Date",
                value=st.session_state.data_manager.get_contract_start_date()
            )
        
        if st.button("💾 Update Contract Settings"):
            st.session_state.data_manager.update_contract_settings(contract_value, contract_start)
            st.success("✅ Contract settings updated!")
            st.rerun()
    
    # P&L Summary
    pl_data = st.session_state.data_manager.get_pl_summary()
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Contract Value", f"${pl_data['contract_value']:,.0f}")
    
    with col2:
        st.metric("Total Costs", f"${pl_data['total_costs']:,.0f}")
    
    with col3:
        profit_loss = pl_data['contract_value'] - pl_data['total_costs']
        st.metric("Profit/Loss", f"${profit_loss:,.0f}")
    
    with col4:
        margin = (profit_loss / pl_data['contract_value'] * 100) if pl_data['contract_value'] > 0 else 0
        st.metric("Margin %", f"{margin:.1f}%")
    
    # P&L Waterfall Chart
    st.markdown("### 💰 Profit & Loss Waterfall")
    
    categories = ['Contract Value', 'Labor Costs', 'ODC', 'Indirect Costs', 'Profit/Loss']
    values = [
        pl_data['contract_value'],
        -pl_data['labor_costs'],
        -pl_data['odc_costs'],
        -pl_data['indirect_costs'],
        profit_loss
    ]
    
    # Create waterfall chart
    fig = go.Figure(go.Waterfall(
        name="P&L Analysis",
        orientation="v",
        measure=["absolute", "relative", "relative", "relative", "total"],
        x=categories,
        textposition="outside",
        text=[f"${v:,.0f}" for v in values],
        y=values,
        connector={"line": {"color": "rgb(63, 63, 63)"}},
        increasing={"marker": {"color": "#2E86AB"}},
        decreasing={"marker": {"color": "#E74C3C"}},
        totals={"marker": {"color": "#27AE60" if profit_loss >= 0 else "#E74C3C"}}
    ))
    
    fig.update_layout(
        title="Contract Profit & Loss Waterfall Analysis",
        showlegend=True,
        height=500
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Cost breakdown over time
    st.markdown("### 📊 Cost Breakdown Over Time")
    
    monthly_pl = st.session_state.data_manager.get_monthly_pl_data()
    
    if not monthly_pl.empty:
        fig = make_subplots(
            rows=2, cols=1,
            subplot_titles=('Monthly Costs', 'Cumulative P&L'),
            specs=[[{"secondary_y": False}], [{"secondary_y": True}]]
        )
        
        # Monthly costs
        fig.add_trace(
            go.Bar(x=monthly_pl['period'], y=monthly_pl['labor'], name='Labor', marker_color='#2E86AB'),
            row=1, col=1
        )
        fig.add_trace(
            go.Bar(x=monthly_pl['period'], y=monthly_pl['odc'], name='ODC', marker_color='#A23B72'),
            row=1, col=1
        )
        fig.add_trace(
            go.Bar(x=monthly_pl['period'], y=monthly_pl['indirect'], name='Indirect', marker_color='#F18F01'),
            row=1, col=1
        )
        
        # Cumulative P&L
        fig.add_trace(
            go.Scatter(x=monthly_pl['period'], y=monthly_pl['cumulative_pl'], name='Cumulative P&L', 
                      line=dict(color='#27AE60', width=3)),
            row=2, col=1
        )
        
        fig.update_layout(height=600, showlegend=True, barmode='stack')
        fig.update_xaxes(title_text="Period", row=2, col=1)
        fig.update_yaxes(title_text="Monthly Costs ($)", row=1, col=1)
        fig.update_yaxes(title_text="Cumulative P&L ($)", row=2, col=1)
        
        st.plotly_chart(fig, use_container_width=True)

def render_cost_import():
    """Render cost data import functionality"""
    
    st.markdown("### 📥 Import Cost Data")
    st.markdown("Import ODC and indirect costs from Excel templates")
    
    # Template Downloads
    st.markdown("#### 📋 Download Templates")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**ODC Template**")
        if st.button("📊 Download ODC Template", type="primary", use_container_width=True):
            odc_template = create_odc_template()
            st.download_button(
                label="⬇️ Download ODC Template",
                data=odc_template,
                file_name=f"ODC_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        st.markdown("**Indirect Costs Template**")
        if st.button("📊 Download Indirect Template", type="secondary", use_container_width=True):
            indirect_template = create_indirect_template()
            st.download_button(
                label="⬇️ Download Indirect Template",
                data=indirect_template,
                file_name=f"Indirect_Costs_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    st.markdown("---")
    
    # File Upload
    st.markdown("#### 📤 Upload Cost Data")
    
    upload_type = st.selectbox("Select data type to import", ["ODC Data", "Indirect Costs"])
    
    uploaded_file = st.file_uploader(
        f"Choose your {upload_type.lower()} file",
        type=['xlsx', 'xls', 'csv'],
        help="Upload an Excel (.xlsx, .xls) or CSV file with cost data"
    )
    
    if uploaded_file is not None:
        try:
            # Process the uploaded file
            if uploaded_file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(uploaded_file)
            else:
                df = pd.read_csv(uploaded_file)
            
            st.success(f"✅ File uploaded successfully! Found {len(df)} records")
            
            # Preview the data
            st.markdown("#### 👀 Data Preview")
            st.dataframe(df.head(10), use_container_width=True)
            
            # Validate and import based on type
            if upload_type == "ODC Data":
                success = import_odc_data(df)
            else:
                success = import_indirect_costs_data(df)
            
            if success:
                st.success("✅ Data imported successfully!")
                st.rerun()
            
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")

def create_odc_template():
    """Create ODC template for download"""
    
    template_data = {
        'category': ['Travel', 'Equipment', 'Software', 'Training', 'Subcontractor'],
        'description': [
            'Business travel to client site',
            'Laptop for new team member',
            'Software license renewal',
            'Technical training course',
            'External consultant services'
        ],
        'vendor': ['Travel Agency', 'Dell', 'Microsoft', 'Training Provider', 'ABC Consulting'],
        'amount': [2500, 1800, 1200, 3000, 15000],
        'date': ['2024-03-15', '2024-03-20', '2024-04-01', '2024-04-15', '2024-05-01'],
        'status': ['Planned', 'Committed', 'Invoiced', 'Planned', 'Committed'],
        'notes': [
            'Round trip for project kickoff',
            'High-performance laptop for development',
            'Annual subscription renewal',
            'Advanced technical certification',
            '3-month consulting engagement'
        ]
    }
    
    df = pd.DataFrame(template_data)
    
    # Create Excel file
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    try:
        df.to_excel(writer, sheet_name='ODC_Data', index=False)
        
        # Style the worksheet
        workbook = writer.book
        worksheet = writer.sheets['ODC_Data']
        
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    finally:
        writer.close()
    
    output.seek(0)
    return output.getvalue()

def create_indirect_template():
    """Create indirect costs template for download"""
    
    # Create monthly periods for 2 years
    periods = []
    costs_data = {'Indirect Costs': ['Fringe', 'Overhead', 'G&A', 'Total Indirect']}
    
    for year in [2024, 2025]:
        for month in range(1, 13):
            period = f"{month}/13-{month+1 if month < 12 else 1}/12/{year if month < 12 else year+1}"
            periods.append(period)
            
            # Sample data
            fringe = 25000 + (month * 1000)
            overhead = 22000 + (month * 900)
            ga = 15000 + (month * 600)
            total = fringe + overhead + ga
            
            costs_data[period] = [fringe, overhead, ga, total]
    
    df = pd.DataFrame(costs_data)
    
    # Create Excel file
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    
    try:
        df.to_excel(writer, sheet_name='Indirect_Costs', index=False)
        
        # Style the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Indirect_Costs']
        
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Style first column
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=1)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    finally:
        writer.close()
    
    output.seek(0)
    return output.getvalue()

def import_odc_data(df):
    """Import ODC data from DataFrame"""
    
    try:
        # Validate required columns
        required_columns = ['category', 'description', 'amount']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            st.error(f"❌ Missing required columns: {', '.join(missing_columns)}")
            return False
        
        # Import each row
        imported_count = 0
        for _, row in df.iterrows():
            if pd.notna(row['description']) and pd.notna(row['amount']):
                odc_data = {
                    'category': str(row.get('category', 'Other')),
                    'description': str(row['description']),
                    'vendor': str(row.get('vendor', '')),
                    'amount': float(row['amount']),
                    'date': pd.to_datetime(row.get('date', datetime.now())).date(),
                    'status': str(row.get('status', 'Planned')),
                    'notes': str(row.get('notes', ''))
                }
                
                st.session_state.data_manager.add_odc_item(odc_data)
                imported_count += 1
        
        st.success(f"✅ Successfully imported {imported_count} ODC items!")
        return True
        
    except Exception as e:
        st.error(f"❌ Import error: {str(e)}")
        return False

def import_indirect_costs_data(df):
    """Import indirect costs data from DataFrame"""
    
    try:
        # Parse the indirect costs structure
        if 'Indirect Costs' not in df.columns:
            st.error("❌ Expected 'Indirect Costs' column not found")
            return False
        
        # Find the rows for Fringe, Overhead, G&A
        cost_types = ['Fringe', 'Overhead', 'G&A']
        imported_count = 0
        
        for col in df.columns:
            if col == 'Indirect Costs' or pd.isna(col) or str(col).startswith('Unnamed'):
                continue
            
            # Parse period from column name
            try:
                # Extract period info - this is flexible to handle different formats
                period_str = str(col)
                
                # For each cost type, get the value
                period_data = {'period': period_str}
                
                for cost_type in cost_types:
                    cost_row = df[df['Indirect Costs'] == cost_type]
                    if not cost_row.empty and col in cost_row.columns:
                        value = cost_row[col].iloc[0]
                        if pd.notna(value) and str(value).replace('.', '').replace('-', '').isdigit():
                            period_data[cost_type.lower()] = float(value)
                        else:
                            period_data[cost_type.lower()] = 0.0
                    else:
                        period_data[cost_type.lower()] = 0.0
                
                # Calculate total
                period_data['total'] = sum([period_data.get('fringe', 0), 
                                          period_data.get('overhead', 0), 
                                          period_data.get('g&a', 0)])
                
                # Save to data manager
                if period_data['total'] > 0:
                    st.session_state.data_manager.update_indirect_costs_period(period_data)
                    imported_count += 1
                    
            except Exception as e:
                continue  # Skip problematic periods
        
        if imported_count > 0:
            st.success(f"✅ Successfully imported indirect costs for {imported_count} periods!")
            return True
        else:
            st.warning("⚠️ No valid indirect cost data found to import")
            return False
        
    except Exception as e:
        st.error(f"❌ Import error: {str(e)}")
        return False